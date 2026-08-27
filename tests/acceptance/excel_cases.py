from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "snooker_vision_user_stories_and_test_cases.xlsx"


@dataclass(frozen=True, slots=True)
class ExcelCase:
    story_id: str
    test_case_id: str
    scenario: str
    expected: str
    test_type: str
    priority: str
    automation: str


def load_p0_cases() -> tuple[ExcelCase, ...]:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    try:
        sheet = workbook["P0测试用例"]
        cases = []
        for row in sheet.iter_rows(min_row=4, max_row=154, values_only=True):
            if not row[5]:
                continue
            cases.append(
                ExcelCase(
                    story_id=str(row[1]),
                    test_case_id=str(row[5]),
                    scenario=str(row[6]),
                    expected=str(row[9]),
                    test_type=str(row[10]),
                    priority=str(row[11]),
                    automation=str(row[12]),
                )
            )
        return tuple(cases)
    finally:
        workbook.close()


def load_p1_cases() -> tuple[ExcelCase, ...]:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    try:
        sheet = workbook["P1测试用例"]
        cases = []
        for row in sheet.iter_rows(min_row=4, max_row=120, values_only=True):
            if not row[5]:
                continue
            cases.append(
                ExcelCase(
                    story_id=str(row[1]),
                    test_case_id=str(row[5]),
                    scenario=str(row[6]),
                    expected=str(row[9]),
                    test_type=str(row[10]),
                    priority=str(row[11]),
                    automation=str(row[12]),
                )
            )
        return tuple(cases)
    finally:
        workbook.close()


def load_traceability_counts() -> dict[str, tuple[int, int, int, int, int]]:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    try:
        sheet = workbook["Story-TC追踪矩阵"]
        return {
            str(row[1]): (int(row[5]), int(row[6]), int(row[7]), int(row[8]), int(row[9]))
            for row in sheet.iter_rows(min_row=4, max_row=18, values_only=True)
        }
    finally:
        workbook.close()


def load_p1_traceability_counts() -> dict[str, tuple[int, int, int, int, int]]:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    try:
        sheet = workbook["Story-TC追踪矩阵"]
        return {
            str(row[1]): (int(row[5]), int(row[6]), int(row[7]), int(row[8]), int(row[9]))
            for row in sheet.iter_rows(min_row=19, max_row=34, values_only=True)
        }
    finally:
        workbook.close()
